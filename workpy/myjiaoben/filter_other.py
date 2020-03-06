# !/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import openpyxl

#设置文件路径
bait_type_path = os.path.dirname(__file__)
workbook=openpyxl.load_workbook(os.path.join(bait_type_path, "excelDate.xlsx"))

# workbook = openpyxl.load_workbook('C:\\Users\\zxc\Desktop\\filterExcel\\excelDate.xlsx')

sheet=workbook.sheetnames
sheet1=workbook.get_sheet_by_name(sheet[0])
rows=sheet1.max_row
cell=sheet1.max_column

sheet_list=[]
for i in range(5,rows+1):
    row_list=[]
    for j in range(1,cell):
        valss=sheet1.cell(i,j).value
        row_list.append(valss)
        if j ==2 :
            if ''.__eq__(str(valss)) or valss == None:
                print('第', i, '行姓名为空')
        if j==3 :
            if ''.__eq__(str(valss)) or valss == None or len(str(valss))!=64:
                print('第', i, '行身份证为空 或身份证密文不足64位')
        if j==4 :
            if ''.__eq__(str(valss)) or valss == None or len(str(valss))!=64:
                print('第', i, '手机号码为空 或 手机密文不足64位')

    # print(row_list)







